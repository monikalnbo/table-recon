#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
recon.py —— 双表关联核对引擎（通用版）
=====================================

场景：两份 Excel，指定一列作关联键（如订单号/快递单号），再自定义若干条核对规则，
引擎输出三类异常并生成标红报告：

  1. 区间不符  —— 规则 type=range：A方是范围（如 0.5-1kg），B方是具体值；B值不落在A方区间内即不符
  2. 项不一致  —— 规则 type=exact：两边必须完全一样（数值用数值相等，文本用归一化后相等）
  3. 单边缺失  —— 关联键只在一边出现

用法（库）:
    from core.recon import compare, export_report
    rules = [
        {"name": "重量", "type": "range", "col_a": "重量区间", "col_b": "实称重量", "tolerance": 0.05,
         "unit_a": "kg", "unit_b": "g"},   # 裸数字时 A方按kg、B方按g理解；带单位的值(800g/0.5kg)自动换算
        {"name": "数量", "type": "exact", "col_a": "数量",   "col_b": "数量"},
    ]
    result = compare("A.xlsx", "B.xlsx", key_a="订单号", key_b="订单号", rules=rules)
    export_report(result, "核对报告.xlsx")

用法（命令行）:
    python3 core/recon.py A.xlsx B.xlsx \
        --key-a 订单号 --key-b 订单号 \
        --rule range:实称重量:重量区间:0.05 \
        --rule exact:数量:数量 \
        -o 核对报告.xlsx --json

仅依赖 openpyxl。区间方向：A方=范围，B方=具体值。单位自动识别并换算（g/kg/公斤/吨/lb…），
区间两侧可各带单位（500g-600g、1-2kg），裸数字默认按 kg，可用规则 unit_a/unit_b 指定。
区间格式：1-2 / 0.5~1.5 / 1至2 / 小于3 / ≥2 / 3以上 / 500g-600g。
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------- 数值/区间解析

UNIT_FACTOR = {
    "mg": 0.000001, "g": 0.001, "克": 0.001,
    "kg": 1.0, "公斤": 1.0, "千克": 1.0,
    "吨": 1000.0, "t": 1000.0,
    "lb": 0.453592, "磅": 0.453592,
    "斤": 0.5,  # 中国电商高频单位（canonical 双修：recon-js 同步）
}
_UNIT_RE = r"(mg|kg|lb|吨|磅|公斤|千克|克|斤|[gt])"


def _norm(text) -> str:
    """NFKC 归一化 + 去空白 + 小写（全角→半角，常见于手输 Excel）"""
    if text is None:
        return ""
    return unicodedata.normalize("NFKC", str(text)).strip().lower().replace(" ", "").replace("～", "~").replace("—", "-").replace("–", "-")


def parse_value(text, default_unit=1.0):
    """解析具体数值（可带单位），统一换算到基准单位 kg；裸数字用 default_unit。
    例: "800g"→0.8, "1.2公斤"→1.2, "0.55"+(default g)→0.00055"""
    s = _norm(text)
    if not s:
        return None
    m = re.search(rf"(-?\d+(?:\.\d+)?)\s*{_UNIT_RE}?", s)
    if not m:
        return None
    try:
        return float(m.group(1)) * UNIT_FACTOR.get(m.group(2), default_unit)
    except ValueError:
        return None


def _unit_of(s: str) -> float:
    if "公斤" in s or "千克" in s or "kg" in s:
        return 1.0
    if "克" in s:
        return 0.001
    for u in ("mg", "lb", "吨", "磅"):
        if u in s:
            return UNIT_FACTOR[u]
    return 1.0  # 缺省按 kg


_NUM = r"(-?\d+(?:\.\d+)?)"


def parse_range(text, default_unit=1.0):
    """解析区间，返回 (low, high) 基准单位 kg；失败返回 None。裸数字用 default_unit。
    单位规则：
      - 区间两侧可各带单位: "500g-600g"→(0.5,0.6)  "1-2kg"→(1,2)  "0.5~1.5公斤"→(0.5,1.5)
      - 只有一侧带单位时作用于整个区间: "500-600g"→(0.5,0.6)
      - 都不带: 按 default_unit
    格式: 1-2 / 0.5~1.5 / 1至2 / 小于3 / 不超过2.5 / ≥2 / >1 / 3以上 / 纯数字视作 [x,x]"""
    s = _norm(text)
    if not s:
        return None
    # a-b / a~b / a至b，两侧可各带单位
    m = re.match(rf"^{_NUM}\s*{_UNIT_RE}?[~\-至]{_NUM}\s*{_UNIT_RE}?", s)
    if m:
        n1, u1, n2, u2 = m.groups()
        f1 = UNIT_FACTOR.get(u1) if u1 else None
        f2 = UNIT_FACTOR.get(u2) if u2 else None
        if f1 is None and f2 is None:
            f1 = f2 = default_unit        # 都裸数字 → 默认单位
        elif f1 is None:
            f1 = f2                        # "500-600g": 尾部单位作用于整个区间
        elif f2 is None:
            f2 = f1                        # "500g-600"
        lo, hi = float(n1) * f1, float(n2) * f2
        return (min(lo, hi), max(lo, hi))
    # 小于/低于/不超过/< x
    m = re.search(rf"(?:小于|低于|不超过|[<≤])\s*{_NUM}\s*{_UNIT_RE}?", s)
    if m:
        n, u = m.groups()
        return (0.0, float(n) * UNIT_FACTOR.get(u, default_unit))
    # 大于/超过/≥/> x
    m = re.search(rf"(?:大于等于|大于|超过|高于|[>≥])\s*{_NUM}\s*{_UNIT_RE}?", s)
    if m:
        n, u = m.groups()
        return (float(n) * UNIT_FACTOR.get(u, default_unit), float("inf"))
    # x以上
    m = re.fullmatch(rf"{_NUM}\s*{_UNIT_RE}?以上", s)
    if m:
        n, u = m.groups()
        return (float(n) * UNIT_FACTOR.get(u, default_unit), float("inf"))
    # 纯数字 → [x, x]
    m = re.fullmatch(rf"{_NUM}\s*{_UNIT_RE}?", s)
    if m:
        n, u = m.groups()
        v = float(n) * UNIT_FACTOR.get(u, default_unit)
        return (v, v)
    return None


def fmt_range(rng) -> str:
    if rng is None:
        return "?"
    lo, hi = rng
    hi_s = "+∞" if hi == float("inf") else f"{hi:g}"
    return f"[{lo:g}, {hi_s}]"


# ---------------------------------------------------------------- 规则核对

def check_range(va, vb, tolerance: float = 0.0, unit_a=1.0, unit_b=1.0):
    """区间核对（有方向）：A方=范围，B方=具体值；B值 ∉ A区间 → 不符。
    unit_a/unit_b：两侧裸数字的默认单位换算系数。所有比较统一在 kg 基准下进行。
    容错：A方解析不出区间而B方可以时自动互换；两边都是纯数值则按容差比大小。
    返回 (ok, 说明)"""
    ra, rb = parse_range(va, unit_a), parse_range(vb, unit_b)
    if ra is not None:                      # 正常路径：A方区间 包含 B方数值
        val = parse_value(vb, unit_b)
        if val is not None:
            return _in_range(ra, val, tolerance, "B", va, vb)
        if rb is not None:                  # B方也是区间 → 相交即可
            lo, hi = max(ra[0], rb[0]), min(ra[1], rb[1])
            if lo - tolerance <= hi + tolerance:
                return True, f"区间相交 {fmt_range(ra)} ∩ {fmt_range(rb)}"
            return False, f"区间不相交 {fmt_range(ra)} ∩ {fmt_range(rb)} = ∅"
        return False, f"B方数值无法解析（原文 {vb!r}）"
    if rb is not None:                      # 容错：方向反了（A值 vs B区间）
        return _in_range(rb, parse_value(va, unit_a), tolerance, "A", vb, va)
    va2, vb2 = parse_value(va, unit_a), parse_value(vb, unit_b)   # 两边都是纯数值
    if va2 is not None and vb2 is not None:
        if abs(va2 - vb2) <= tolerance + 1e-9:
            return True, f"数值相等 {va2:g} = {vb2:g}"
        return False, f"数值不等 {va2:g} ≠ {vb2:g}"
    return False, "两边都无法解析为数值/区间"


def _in_range(rng, val, tolerance, val_side, range_text, val_text):
    if val is None:
        return False, f"{val_side}方数值无法解析（原文 {val_text!r}）"
    lo, hi = rng
    if lo - tolerance <= val <= hi + tolerance:
        return True, f"{val:g} ∈ {fmt_range(rng)}"
    return False, f"{val:g} ∉ {fmt_range(rng)}（区间原文 {range_text!r}）"


def check_exact(va, vb):
    """精确核对：数值按数值比（1.0 == 1），文本按归一化字符串比。返回 (ok, 说明)"""
    na, nb = parse_value(va), parse_value(vb)
    if na is not None and nb is not None:
        if abs(na - nb) < 1e-9:
            return True, "数值相等"
        return False, f"数值不等: {va} ≠ {vb}"
    sa, sb = _norm(va), _norm(vb)
    if sa == sb:
        return True, "文本一致"
    return False, f"文本不一致: {va!r} ≠ {vb!r}"


# ---------------------------------------------------------------- 表格读取

def read_sheet(path, sheet=None):
    """读取 Excel 第一个 sheet（或指定 sheet），返回 (headers, rows[list of dict])"""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    rows = ws.iter_rows(values_only=True)
    headers = [str(h).strip() if h is not None else "" for h in next(rows, [])]
    data = []
    for r in rows:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        data.append({headers[i]: r[i] for i in range(len(headers)) if i < len(r)})
    wb.close()
    return headers, data


def _keyof(row, col):
    v = row.get(col)
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        v = int(v)  # 订单号 123 vs "123" 统一
    return _norm(v)


# ---------------------------------------------------------------- 主核对

def compare(file_a, file_b, key_a, key_b, rules, tolerance=0.0, sheet_a=None, sheet_b=None):
    """
    核对两份表。rules: [{name,type:range|exact,col_a,col_b,tolerance?}, ...]
    返回结果 dict（可直接 json.dumps，也是 export_report 的输入）。
    """
    ha, ra = read_sheet(file_a, sheet_a)
    hb, rb = read_sheet(file_b, sheet_b)
    for col, side, headers in [(key_a, "A", ha), (key_b, "B", hb)]:
        if col not in headers:
            raise ValueError(f"{side}方关联列 {col!r} 不存在，可选列: {headers}")
    for i, rule in enumerate(rules):
        for side, headers, col in [("A", ha, rule["col_a"]), ("B", hb, rule["col_b"])]:
            if col not in headers:
                raise ValueError(f"规则{i + 1}（{rule.get('name', rule['type'])}）{side}方列 {col!r} 不存在，可选列: {headers}")

    map_a = {_keyof(r, key_a): r for r in ra if _keyof(r, key_a)}
    map_b = {_keyof(r, key_b): r for r in rb if _keyof(r, key_b)}
    keys_a, keys_b = set(map_a), set(map_b)

    range_miss, exact_miss, only_a, only_b, ok_count = [], [], [], [], 0
    for k in sorted(keys_a & keys_b):
        problems = []
        for rule in rules:
            va, vb = map_a[k].get(rule["col_a"]), map_b[k].get(rule["col_b"])
            if rule["type"] == "range":
                ok, why = check_range(va, vb, rule.get("tolerance", tolerance),
                                      _unit_factor(rule.get("unit_a")), _unit_factor(rule.get("unit_b")))
            else:
                ok, why = check_exact(va, vb)
            if not ok:
                item = {"key": k, "rule": rule.get("name", rule["type"]), "type": rule["type"],
                        "col_a": rule["col_a"], "col_b": rule["col_b"], "a": _plain(va), "b": _plain(vb), "reason": why}
                (range_miss if rule["type"] == "range" else exact_miss).append(item)
                problems.append(item)
        if not problems:
            ok_count += 1

    for k in sorted(keys_a - keys_b):
        only_a.append({"key": k, "row": _row_plain(map_a[k])})
    for k in sorted(keys_b - keys_a):
        only_b.append({"key": k, "row": _row_plain(map_b[k])})

    return {
        "files": {"a": str(file_a), "b": str(file_b)},
        "keys": {"a": key_a, "b": key_b},
        "rules": rules,
        "summary": {
            "A方记录": len(ra), "B方记录": len(rb),
            "匹配": ok_count,
            "区间不符": len(range_miss),
            "项不一致": len(exact_miss),
            "仅A方有": len(only_a),
            "仅B方有": len(only_b),
        },
        "range_mismatch": range_miss,
        "exact_mismatch": exact_miss,
        "only_in_a": only_a,
        "only_in_b": only_b,
        "_raw": {"headers_a": ha, "headers_b": hb, "map_a": map_a, "map_b": map_b,
                 "key_a": key_a, "key_b": key_b, "rows_a": ra, "rows_b": rb},
    }


def _unit_factor(spec):
    """单位说明 → 换算系数。支持单位名（kg/g/公斤/吨/lb…）或直接给系数（如 0.001）"""
    if not spec:
        return 1.0
    if spec in UNIT_FACTOR:
        return UNIT_FACTOR[spec]
    try:
        return float(spec)
    except (TypeError, ValueError):
        return 1.0


def _plain(v):
    return None if v is None else (int(v) if isinstance(v, float) and v.is_integer() else v)


def _row_plain(row):
    return {k: _plain(v) for k, v in row.items()}


# ---------------------------------------------------------------- 报告导出

_RED = PatternFill("solid", fgColor="FFC7CE")
_RED_FONT = Font(color="9C0006")
_HEAD = PatternFill("solid", fgColor="DDEBF7")


def export_report(result, out_path):
    """生成标红核对报告：总览 / 区间不符 / 项不一致 / 仅A方 / 仅B方 / A/B方标注"""
    raw = result["_raw"]
    wb = Workbook()
    wb.remove(wb.active)

    def sheet(title, headers, rows, red_when=None):
        ws = wb.create_sheet(title)
        ws.append(headers)
        for c in ws[1]:
            c.fill, c.font = _HEAD, Font(bold=True)
        for r in rows:
            ws.append(r)
            if red_when and red_when(r):
                for c in ws[ws.max_row]:
                    c.fill, c.font = _RED, _RED_FONT
        for i, h in enumerate(headers, 1):
            w = max(len(str(h)), *(len(str(r[i - 1])) if i - 1 < len(r) and r[i - 1] is not None else 0 for r in rows)) if rows else len(str(h))
            ws.column_dimensions[get_column_letter(i)].width = min(max(w * 1.8, 10), 45)
        ws.freeze_panes = "A2"
        return ws

    s = result["summary"]
    overview = [
        ["A方文件", result["files"]["a"]], ["B方文件", result["files"]["b"]],
        ["关联列", f"A:{result['keys']['a']} ↔ B:{result['keys']['b']}"],
    ] + [[k, v] for k, v in s.items()] + [
        ["核对规则", "; ".join(f"{r.get('name', r['type'])}({r['type']}: {r['col_a']}↔{r['col_b']})" for r in result["rules"])],
    ]
    sheet("总览", ["项目", "数量"], overview)

    sheet("区间不符", ["关联键", "规则", "A方值", "B方值", "说明"],
          [[m["key"], m["rule"], m["a"], m["b"], m["reason"]] for m in result["range_mismatch"]])
    sheet("项不一致", ["关联键", "规则", "A方值", "B方值", "说明"],
          [[m["key"], m["rule"], m["a"], m["b"], m["reason"]] for m in result["exact_mismatch"]])
    sheet("仅A方有", ["关联键"] + raw["headers_a"],
          [[o["key"]] + [_plain(o["row"].get(h)) for h in raw["headers_a"]] for o in result["only_in_a"]])
    sheet("仅B方有", ["关联键"] + raw["headers_b"],
          [[o["key"]] + [_plain(o["row"].get(h)) for h in raw["headers_b"]] for o in result["only_in_b"]])

    # A/B 方标注：原表副本，异常行整行标红
    bad_a = {m["key"] for m in result["range_mismatch"] + result["exact_mismatch"]} | {o["key"] for o in result["only_in_a"]}
    bad_b = {m["key"] for m in result["range_mismatch"] + result["exact_mismatch"]} | {o["key"] for o in result["only_in_b"]}
    sheet("A方标注", raw["headers_a"], [[_plain(r.get(h)) for h in raw["headers_a"]] for r in raw["rows_a"]],
          red_when=lambda r: _keyof(dict(zip(raw["headers_a"], r)), raw["key_a"]) in bad_a)
    sheet("B方标注", raw["headers_b"], [[_plain(r.get(h)) for h in raw["headers_b"]] for r in raw["rows_b"]],
          red_when=lambda r: _keyof(dict(zip(raw["headers_b"], r)), raw["key_b"]) in bad_b)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------- CLI

def _parse_rule(spec: str):
    """'range:A列:B列[:容差[:A单位[:B单位]]]' —— range 规则A列=区间、B列=具体值；
    例: range:重量区间:实称重量:0.05:kg:g  （A方裸数字按kg、B方裸数字按g）
    'exact:数量:数量'"""
    parts = spec.split(":")
    t = parts[0].lower()
    alias = {"区间": "range", "精确": "exact"}
    t = alias.get(t, t)
    if t not in ("range", "exact") or len(parts) < 3:
        raise ValueError(f"规则格式: range:A列:B列[:容差] 或 exact:A列:B列 —— 收到 {spec!r}")
    rule = {"type": t, "col_a": parts[1], "col_b": parts[2], "name": f"{parts[1]}↔{parts[2]}"}
    if len(parts) > 3 and parts[3]:
        rule["tolerance"] = float(parts[3])
    if len(parts) > 4 and parts[4]:
        rule["unit_a"] = parts[4]
    if len(parts) > 5 and parts[5]:
        rule["unit_b"] = parts[5]
    return rule


def main():
    ap = argparse.ArgumentParser(description="双表关联核对：区间比对 / 精确比对 / 单边缺失")
    ap.add_argument("file_a"), ap.add_argument("file_b")
    ap.add_argument("--key-a", required=True, help="A方关联列名，如 订单号")
    ap.add_argument("--key-b", required=True, help="B方关联列名")
    ap.add_argument("--rule", action="append", default=[], metavar="TYPE:A列:B列[:容差[:A单位[:B单位]]]",
                    help="可重复。range|区间: A列范围 包含 B列值; exact|精确: 两列完全一致。单位如 kg/g/公斤/吨/lb")
    ap.add_argument("--tolerance", type=float, default=0.0, help="全局容差（区间规则未单独指定时生效）")
    ap.add_argument("--sheet-a"), ap.add_argument("--sheet-b")
    ap.add_argument("-o", "--output", default="核对报告.xlsx")
    ap.add_argument("--json", action="store_true", help="结果打印为 JSON（供 Agent/MCP 使用）")
    args = ap.parse_args()

    rules = [_parse_rule(r) for r in args.rule]
    result = compare(args.file_a, args.file_b, args.key_a, args.key_b, rules,
                     tolerance=args.tolerance, sheet_a=args.sheet_a, sheet_b=args.sheet_b)
    export_report(result, args.output)

    slim = {k: v for k, v in result.items() if not k.startswith("_")}
    if args.json:
        print(json.dumps(slim, ensure_ascii=False, indent=2))
    else:
        s = result["summary"]
        print(f"✔ 核对完成，报告已生成: {args.output}")
        for k, v in s.items():
            print(f"  {k}: {v}")
        for title, items in [("区间不符", result["range_mismatch"]), ("项不一致", result["exact_mismatch"])]:
            for m in items:
                print(f"  [{title}] {m['key']}: {m['a']} vs {m['b']} —— {m['reason']}")
        for title, items in [("仅A方有", result["only_in_a"]), ("仅B方有", result["only_in_b"])]:
            for o in items:
                print(f"  [{title}] {o['key']}")


if __name__ == "__main__":
    main()
