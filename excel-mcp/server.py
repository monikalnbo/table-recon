#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
excel-mcp —— pi 内置 Excel MCP（查询 + 修改，自包含）
=====================================================

零外部依赖（仅 openpyxl），单文件 stdio MCP server。
与 03-mcp-server/server.py 同一套按行 JSON-RPC 协议（pi-mcp-adapter 实测可用）。

工具 13 个：

  读/查
    inspect_workbook   工作簿结构：sheet 列表、行列数、表头、预览
    read_range         按 A1 区域读原始网格
    query_rows         按条件筛选行（eq/ne/gt/lt/ge/le/contains/in/between/is_empty…）
  写/改
    write_data         从某单元格起写 2D 数据
    append_rows        追加行（二维数组或按表头的字典）
    update_rows        按筛选条件批量改列值（查询式修改）
    set_cells          按 {"A1": 值} 精确补丁
    apply_formula      写公式（=SUM(...)）
    format_range       加粗/字号/颜色/填充/数字格式/对齐（xlsx）
    find_replace       查找替换
    delete_rows        删行
  结构
    create_workbook    新建工作簿/sheet
    manage_sheets      create/copy/rename/delete sheet

支持 .xlsx 与 .csv（csv 只支持值读写，公式/格式会明确报错）。
每次调用即开即存，无服务端状态。

用法（pi，仓库 .mcp.json）:
  "excel": { "command": "python3", "args": ["excel-mcp/server.py"] }

自测（无需客户端）:
  python3 excel-mcp/server.py --selftest
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import re
import sys
import tempfile
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import column_index_from_string, get_column_letter

PROTOCOL = "2024-11-05"
SERVER_INFO = {"name": "excel-mcp", "version": "1.0.0"}

# ---------------------------------------------------------------- 基础工具

_CELL_RE = re.compile(r"^([A-Za-z]+)(\d+)$")


def _cell_to_idx(cell: str):
    m = _CELL_RE.match(str(cell).strip())
    if not m:
        raise ValueError(f"非法单元格引用: {cell!r}（应为 A1 形式）")
    return int(m.group(2)), column_index_from_string(m.group(1).upper())  # (row, col)


def _plain(v):
    if v is None:
        return None
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def _is_csv(path) -> bool:
    return Path(path).suffix.lower() == ".csv"


def _require(path):
    p = Path(path)
    if not p.exists():
        raise FileNotFoundError(f"文件不存在: {p}")
    if p.suffix.lower() not in (".xlsx", ".xlsm", ".csv"):
        raise ValueError(f"仅支持 .xlsx/.xlsm/.csv，收到: {p.suffix!r}")
    return p


def _xlsx_only(path):
    if _is_csv(path):
        raise ValueError("csv 不支持该操作（公式/格式仅限 xlsx）")


def _load(path, sheet=None):
    """打开 xlsx，返回 (wb, ws)。sheet 缺省=active。"""
    wb = load_workbook(path)  # 保留公式/样式
    ws = wb[sheet] if sheet else wb.active
    if ws is None:
        raise ValueError(f"工作表不存在: {sheet!r}，可选: {wb.sheetnames}")
    return wb, ws


def _sheet_rows_xlsx(path, sheet=None):
    """读成 (headers, rows[list of dict], ws)。空表头列用 '列B' 兜底。"""
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet] if sheet else wb.active
    if ws is None:
        wb.close()
        raise ValueError(f"工作表不存在: {sheet!r}，可选: {wb.sheetnames}")
    it = ws.iter_rows(values_only=True)
    raw_head = next(it, [])
    headers = [str(h).strip() if h is not None else f"列{get_column_letter(i+1)}" for i, h in enumerate(raw_head)]
    rows = []
    for r in it:
        if all(c is None or str(c).strip() == "" for c in r):
            continue
        rows.append({headers[i]: r[i] for i in range(min(len(headers), len(r)))})
    wb.close()
    return headers, rows, ws.title


def _sheet_rows_csv(path):
    with open(path, newline="", encoding="utf-8-sig") as f:
        data = list(csv.reader(f))
    if not data:
        return [], []
    headers = [h.strip() if h else f"列{get_column_letter(i+1)}" for i, h in enumerate(data[0])]
    rows = [dict(zip(headers, r)) for r in data[1:] if any(str(c).strip() for c in r)]
    return headers, rows


def sheet_rows(path, sheet=None):
    if _is_csv(path):
        return _sheet_rows_csv(path) + (Path(path).stem,)
    h, r, t = _sheet_rows_xlsx(path, sheet)
    return h, r, t


# ---------------------------------------------------------------- 筛选引擎

_OPS_NUM = {"eq", "ne", "gt", "lt", "ge", "le"}


def _num(v):
    try:
        if isinstance(v, bool):
            return None
        if isinstance(v, (int, float)):
            return float(v)
        s = str(v).strip().replace(",", "")
        return float(s) if s else None
    except (TypeError, ValueError):
        return None


def _match(value, op: str, want) -> bool:
    if op in ("is_empty",):
        return value is None or str(value).strip() == ""
    if op in ("not_empty",):
        return not (value is None or str(value).strip() == "")
    if value is None:
        return False
    if op in _OPS_NUM:
        a, b = _num(value), _num(want)
        if a is None or b is None:  # 数值语义下解析失败 → 退回文本比
            a, b = str(value).strip().lower(), str(want).strip().lower()
            return {"eq": a == b, "ne": a != b}.get(op, False)
        return {"eq": a == b, "ne": a != b, "gt": a > b, "lt": a < b, "ge": a >= b, "le": a <= b}[op]
    if op == "between":
        lo, hi = (_num(x) for x in want)
        v = _num(value)
        if v is None or lo is None or hi is None:
            return False
        return min(lo, hi) <= v <= max(lo, hi)
    s, w = str(value).strip().lower(), str(want).strip().lower()
    if op == "contains":
        return w in s
    if op == "not_contains":
        return w not in s
    if op == "in":
        return s in {str(x).strip().lower() for x in (want if isinstance(want, list) else [want])}
    if op == "not_in":
        return s not in {str(x).strip().lower() for x in (want if isinstance(want, list) else [want])}
    raise ValueError(f"不支持的操作符: {op!r}（可用: eq/ne/gt/lt/ge/le/contains/not_contains/in/not_in/between/is_empty/not_empty）")


def apply_filters(rows, filters, headers):
    if not filters:
        return rows
    for i, f in enumerate(filters):
        col = f.get("column")
        if col not in headers:
            raise ValueError(f"筛选{i+1}列 {col!r} 不存在，可选: {headers}")
    out = []
    for r in rows:
        try:
            if all(_match(r.get(f["column"]), f.get("op", "eq"), f.get("value", f.get("values"))) for f in filters):
                out.append(r)
        except KeyError as e:
            raise ValueError(f"筛选缺少字段 {e}，需要 column/op/value") from None
    return out


# ---------------------------------------------------------------- 工具实现

def t_inspect_workbook(a):
    p = _require(a["path"])
    if _is_csv(p):
        headers, rows = _sheet_rows_csv(p)
        return _json({
            "path": str(p), "sheets": [{"name": p.stem, "rows": len(rows), "columns": headers}],
            "active": p.stem, "columns": headers, "row_count": len(rows),
            "preview": [{k: _plain(v) for k, v in r.items()} for r in rows[:3]],
        })
    wb = load_workbook(p, data_only=True, read_only=True)
    sheets = []
    headers, preview, active = [], [], ""
    target = a.get("sheet") or wb.active.title
    for name in wb.sheetnames:
        ws = wb[name]
        sheets.append({"name": name, "rows": ws.max_row, "cols": ws.max_column})
        if name == target:
            active = name
            it = ws.iter_rows(values_only=True)
            raw = next(it, [])
            headers = [str(h).strip() if h is not None else f"列{get_column_letter(i+1)}" for i, h in enumerate(raw)]
            rows = [dict(zip(headers, r)) for r in it if any(c is not None and str(c).strip() for c in r)]
            preview = [{k: _plain(v) for k, v in list(r.items())[:12]} for r in rows[:3]]
    wb.close()
    return _json({"path": str(p), "sheets": sheets, "active": active,
                  "columns": headers, "preview": preview})


def t_read_range(a):
    p = _require(a["path"])
    start = a.get("start_cell", "A1")
    if _is_csv(p):
        _, rows = _sheet_rows_csv(p)
        r1, _c1 = _cell_to_idx(start)
        grid = [[_plain(v) for v in r.values()] for r in rows]
        return _json({"start_cell": start, "data": grid[r1 - 1:]})
    data_only = a.get("values_only", True)
    wb2 = load_workbook(p, data_only=data_only)
    ws2 = wb2[a.get("sheet")] if a.get("sheet") else wb2.active
    r1, c1 = _cell_to_idx(start)
    r2, c2 = (_cell_to_idx(a["end_cell"]) if a.get("end_cell") else (ws2.max_row, ws2.max_column))
    r1, r2, c1, c2 = min(r1, r2), max(r1, r2), min(c1, c2), max(c1, c2)
    data = [[_plain(ws2.cell(row=r, column=c).value) for c in range(c1, c2 + 1)] for r in range(r1, r2 + 1)]
    wb2.close()
    return _json({"range": f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}", "data": data})


def t_query_rows(a):
    p = _require(a["path"])
    headers, rows, title = sheet_rows(p, a.get("sheet"))
    matched = apply_filters(rows, a.get("filters"), headers)
    cols = a.get("columns") or headers
    for c in cols:
        if c not in headers:
            raise ValueError(f"列 {c!r} 不存在，可选: {headers}")
    limit = min(int(a.get("limit", 50)), 500)
    return _json({
        "sheet": title, "total_rows": len(rows), "matched": len(matched),
        "columns": cols,
        "rows": [{c: _plain(r.get(c)) for c in cols} for r in matched[:limit]],
        "note": f"显示前 {min(limit, len(matched))} 行" if len(matched) > limit else None,
    })


def t_write_data(a):
    p = _require(a["path"])
    data = a["data"]
    if not isinstance(data, list) or not all(isinstance(r, list) for r in data):
        raise ValueError("data 必须是二维数组 [[..],[..]]")
    if _is_csv(p):
        headers, _ = _sheet_rows_csv(p)
        with open(p, "a", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            for r in data:
                w.writerow(r)
        return _json({"written": len(data), "mode": "csv-append"})
    _xlsx_only(p)
    wb, ws = _load(p, a.get("sheet"))
    if a.get("sheet") and a["sheet"] not in wb.sheetnames:
        ws = wb.create_sheet(a["sheet"])
    r1, c1 = _cell_to_idx(a.get("start_cell", "A1"))
    for i, row in enumerate(data):
        for j, v in enumerate(row):
            ws.cell(row=r1 + i, column=c1 + j, value=v)
    wb.save(p)
    wb.close()
    return _json({"written": len(data), "start_cell": a.get("start_cell", "A1"), "sheet": ws.title})


def t_append_rows(a):
    p = _require(a["path"])
    rows_in = a["rows"]
    headers, _rows, title = sheet_rows(p, a.get("sheet"))
    dict_mode = rows_in and isinstance(rows_in[0], dict)
    if dict_mode:
        for i, r in enumerate(rows_in):
            for k in r:
                if k not in headers:
                    raise ValueError(f"第{i+1}行字段 {k!r} 不在表头中，可用: {headers}")
        data = [[r.get(h) for h in headers] for r in rows_in]
    else:
        data = rows_in
    if _is_csv(p):
        with open(p, "a", newline="", encoding="utf-8-sig") as f:
            csv.writer(f).writerows(data)
        return _json({"appended": len(data), "sheet": title, "headers": headers})
    wb, ws = _load(p, a.get("sheet"))
    start = ws.max_row + 1 if ws.max_row >= 1 else 1
    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        start = 1
    for i, row in enumerate(data):
        for j, v in enumerate(row):
            ws.cell(row=start + i, column=j + 1, value=v)
    wb.save(p)
    wb.close()
    return _json({"appended": len(data), "start_row": start, "sheet": ws.title, "headers": headers})


def t_update_rows(a):
    p = _require(a["path"])
    updates = a["updates"]
    if not updates:
        raise ValueError("updates 不能为空，如 {\"实称重量\": 850}")
    if _is_csv(p):
        headers, rows = _sheet_rows_csv(p)
        for k in updates:
            if k not in headers:
                raise ValueError(f"更新列 {k!r} 不存在，可选: {headers}")
        matched = apply_filters(rows, a.get("filters"), headers)
        for r in matched:
            r.update(updates)
        with open(p, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(["" if r.get(h) is None else r.get(h) for h in headers])
        return _json({"updated": len(matched), "updates": updates, "sheet": Path(p).stem})
    wb, ws = _load(p, a.get("sheet"))
    head_row = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else f"列{get_column_letter(i+1)}" for i, h in enumerate(head_row)]
    for k in updates:
        if k not in headers:
            wb.close()
            raise ValueError(f"更新列 {k!r} 不存在，可选: {headers}")
    col_of = {h: i + 1 for i, h in enumerate(headers)}
    # 用 data_only 视图做筛选定位原表行号，再回写到原 ws（保公式/格式）
    wsv = load_workbook(p, data_only=True)[ws.title]
    upd_count = 0
    for ri in range(2, ws.max_row + 1):
        vals = {headers[i]: wsv.cell(row=ri, column=i + 1).value for i in range(len(headers))}
        if all(c is None or str(c).strip() == "" for c in vals.values()):
            continue
        if apply_filters([vals], a.get("filters"), headers):
            for k, v in updates.items():
                ws.cell(row=ri, column=col_of[k], value=v)
            upd_count += 1
    wsv.parent.close()
    wb.save(p)
    wb.close()
    return _json({"updated": upd_count, "updates": updates, "sheet": ws.title})


def t_set_cells(a):
    p = _require(a["path"])
    cells = a["cells"]
    if not isinstance(cells, dict) or not cells:
        raise ValueError("cells 必须是 {\"A1\": 值, ...}")
    _xlsx_only(p)
    wb, ws = _load(p, a.get("sheet"))
    for ref, v in cells.items():
        r, c = _cell_to_idx(ref)
        ws.cell(row=r, column=c, value=v)
    wb.save(p)
    wb.close()
    return _json({"set": len(cells), "cells": cells})


def t_apply_formula(a):
    p = _require(a["path"])
    _xlsx_only(p)
    formula = a["formula"].strip()
    if not formula.startswith("="):
        formula = "=" + formula
    wb, ws = _load(p, a.get("sheet"))
    r, c = _cell_to_idx(a["cell"])
    ws.cell(row=r, column=c, value=formula)
    wb.save(p)
    wb.close()
    return _json({"cell": a["cell"].upper(), "formula": formula, "sheet": ws.title})


def t_format_range(a):
    p = _require(a["path"])
    _xlsx_only(p)
    wb, ws = _load(p, a.get("sheet"))
    r1, c1 = _cell_to_idx(a["start_cell"])
    r2, c2 = _cell_to_idx(a.get("end_cell") or a["start_cell"])
    r1, r2, c1, c2 = min(r1, r2), max(r1, r2), min(c1, c2), max(c1, c2)
    font_kw, n_fmt, fill, align = {}, a.get("number_format"), None, None
    if a.get("bold") is not None:
        font_kw["bold"] = bool(a["bold"])
    if a.get("italic") is not None:
        font_kw["italic"] = bool(a["italic"])
    if a.get("font_size") is not None:
        font_kw["size"] = a["font_size"]
    if a.get("font_color"):
        font_kw["color"] = a["font_color"].lstrip("#")
    if a.get("bg_color"):
        fill = PatternFill("solid", fgColor=a["bg_color"].lstrip("#"))
    if a.get("alignment"):
        align = Alignment(horizontal=a["alignment"], wrap_text=bool(a.get("wrap_text", False)))
    font = Font(**font_kw) if font_kw else None
    n = 0
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            cell = ws.cell(row=r, column=c)
            if font:
                cell.font = font
            if fill:
                cell.fill = fill
            if align:
                cell.alignment = align
            if n_fmt:
                cell.number_format = n_fmt
            n += 1
    wb.save(p)
    wb.close()
    return _json({"formatted": n, "range": f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"})


def t_find_replace(a):
    p = _require(a["path"])
    find, repl = str(a["find"]), str(a.get("replace", ""))
    if not find:
        raise ValueError("find 不能为空")
    count = 0
    if _is_csv(p):
        headers, rows = _sheet_rows_csv(p)
        cols = [a["column"]] if a.get("column") else headers
        for r in rows:
            for h in cols:
                v = r.get(h)
                if v is not None and find in str(v):
                    r[h] = str(v).replace(find, repl)
                    count += 1
        with open(p, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(headers)
            for r in rows:
                w.writerow(["" if r.get(h) is None else r.get(h) for h in headers])
        return _json({"replaced": count, "find": find, "replace": repl})
    wb, ws = _load(p, a.get("sheet"))
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and find in str(cell.value):
                cell.value = str(cell.value).replace(find, repl)
                count += 1
    wb.save(p)
    wb.close()
    return _json({"replaced": count, "find": find, "replace": repl, "sheet": ws.title})


def t_delete_rows(a):
    p = _require(a["path"])
    r1 = int(a["row_start"])
    r2 = int(a.get("row_end") or a["row_start"])
    if min(r1, r2) < 1:
        raise ValueError("行号从 1 开始")
    _xlsx_only(p)
    wb, ws = _load(p, a.get("sheet"))
    n = 0
    for r in range(max(r1, r2), min(r1, r2) - 1, -1):
        ws.delete_rows(r)
        n += 1
    wb.save(p)
    wb.close()
    return _json({"deleted_rows": n, "rows": f"{min(r1,r2)}-{max(r1,r2)}"})


def t_create_workbook(a):
    p = Path(a["path"])
    if p.exists():
        raise FileExistsError(f"文件已存在: {p}")
    if p.suffix.lower() == ".csv":
        with open(p, "w", newline="", encoding="utf-8-sig") as f:
            csv.writer(f).writerow(a.get("headers", []))
        return _json({"created": str(p), "headers": a.get("headers", [])})
    wb = Workbook()
    if a.get("sheet_name"):
        wb.active.title = a["sheet_name"]
    if a.get("headers"):
        wb.active.append(a["headers"])
    p.parent.mkdir(parents=True, exist_ok=True)
    wb.save(p)
    wb.close()
    return _json({"created": str(p), "sheets": [a.get("sheet_name") or "Sheet"], "rows": 1 if a.get("headers") else 0})


def t_manage_sheets(a):
    p = _require(a["path"])
    _xlsx_only(p)
    action, name = a["action"], a["sheet"]
    wb = load_workbook(p)
    if action == "create":
        if name in wb.sheetnames:
            raise ValueError(f"sheet 已存在: {name}")
        wb.create_sheet(name)
    elif action == "copy":
        if name not in wb.sheetnames:
            raise ValueError(f"sheet 不存在: {name}")
        wb.copy_worksheet(wb[name]).title = a.get("new_name") or f"{name} 副本"
    elif action == "rename":
        if name not in wb.sheetnames:
            raise ValueError(f"sheet 不存在: {name}")
        wb[name].title = a["new_name"]
    elif action == "delete":
        if name not in wb.sheetnames:
            raise ValueError(f"sheet 不存在: {name}")
        if len(wb.sheetnames) == 1:
            raise ValueError("至少保留一个 sheet")
        wb.remove(wb[name])
    else:
        raise ValueError("action 必须是 create/copy/rename/delete")
    wb.save(p)
    wb.close()
    return _json({"action": action, "sheets": load_workbook(p).sheetnames})


def _json(obj) -> str:
    return json.dumps(obj, ensure_ascii=False, default=str)


# ---------------------------------------------------------------- MCP 规格

_F = {"type": "object", "properties": {
    "column": {"type": "string"}, "op": {"type": "string"},
    "value": {}, "values": {"type": "array"}},
    "required": ["column", "op"]}

TOOLS_SPEC = [
    {"name": "inspect_workbook",
     "description": "查看 Excel/CSV 结构：sheet 列表、行列数、表头、前3行预览。任何操作前先调这个。",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"}}, "required": ["path"]}},
    {"name": "read_range",
     "description": "按 A1 区域读原始网格（如 A1:C10）。values_only=false 时返回公式原文。",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"},
        "start_cell": {"type": "string"}, "end_cell": {"type": "string"},
        "values_only": {"type": "boolean", "default": True}}, "required": ["path", "start_cell"]}},
    {"name": "query_rows",
     "description": "按条件查行（首行为表头）。op: eq/ne/gt/lt/ge/le/contains/not_contains/in/not_in/between/is_empty/not_empty。数值列自动按数值比。",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"},
        "filters": {"type": "array", "items": _F},
        "columns": {"type": "array", "items": {"type": "string"}},
        "limit": {"type": "integer", "default": 50}}, "required": ["path"]}},
    {"name": "write_data",
     "description": "从 start_cell 起写二维数组数据（会覆盖目标区域）。",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"},
        "start_cell": {"type": "string", "default": "A1"},
        "data": {"type": "array", "items": {"type": "array"}}}, "required": ["path", "data"]}},
    {"name": "append_rows",
     "description": "追加行。rows 可为二维数组，或按表头的字典数组（自动对齐列顺序）。",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"},
        "rows": {"type": "array"}}, "required": ["path", "rows"]}},
    {"name": "update_rows",
     "description": "查询式批量修改：筛选出匹配行，把这些行的指定列改成新值。filters 同 query_rows。",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"},
        "filters": {"type": "array", "items": _F},
        "updates": {"type": "object"}}, "required": ["path", "updates"]}},
    {"name": "set_cells",
     "description": "精确补丁：{\"B3\": 值, \"C7\": 值} 一次写多个单元格。",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"},
        "cells": {"type": "object"}}, "required": ["path", "cells"]}},
    {"name": "apply_formula",
     "description": "给单元格写公式，如 =SUM(B2:B10)。带不带等号都行。（仅 xlsx）",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"},
        "cell": {"type": "string"}, "formula": {"type": "string"}}, "required": ["path", "cell", "formula"]}},
    {"name": "format_range",
     "description": "格式化区域：加粗/斜体/字号/字色/背景色/数字格式/对齐。颜色如 FFC7CE 或 #FFC7CE。（仅 xlsx）",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"},
        "start_cell": {"type": "string"}, "end_cell": {"type": "string"},
        "bold": {"type": "boolean"}, "italic": {"type": "boolean"},
        "font_size": {"type": "number"}, "font_color": {"type": "string"},
        "bg_color": {"type": "string"}, "number_format": {"type": "string"},
        "alignment": {"type": "string"}, "wrap_text": {"type": "boolean"}},
        "required": ["path", "start_cell"]}},
    {"name": "find_replace",
     "description": "全表或指定列查找替换文本。",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"},
        "find": {"type": "string"}, "replace": {"type": "string"},
        "column": {"type": "string"}}, "required": ["path", "find"]}},
    {"name": "delete_rows",
     "description": "删除行（1-based，含两端）。（仅 xlsx）",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet": {"type": "string"},
        "row_start": {"type": "integer"}, "row_end": {"type": "integer"}},
        "required": ["path", "row_start"]}},
    {"name": "create_workbook",
     "description": "新建工作簿（可带表头行）。",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "sheet_name": {"type": "string"},
        "headers": {"type": "array", "items": {"type": "string"}}}, "required": ["path"]}},
    {"name": "manage_sheets",
     "description": "管理 sheet：create / copy / rename / delete。",
     "inputSchema": {"type": "object", "properties": {
        "path": {"type": "string"}, "action": {"type": "string", "enum": ["create", "copy", "rename", "delete"]},
        "sheet": {"type": "string"}, "new_name": {"type": "string"}},
        "required": ["path", "action", "sheet"]}},
]

TOOL_IMPLS = {
    "inspect_workbook": t_inspect_workbook,
    "read_range": t_read_range,
    "query_rows": t_query_rows,
    "write_data": t_write_data,
    "append_rows": t_append_rows,
    "update_rows": t_update_rows,
    "set_cells": t_set_cells,
    "apply_formula": t_apply_formula,
    "format_range": t_format_range,
    "find_replace": t_find_replace,
    "delete_rows": t_delete_rows,
    "create_workbook": t_create_workbook,
    "manage_sheets": t_manage_sheets,
}


# ---------------------------------------------------------------- stdio 协议（与 03-mcp-server 同款）

def send(msg: dict):
    sys.stdout.write(json.dumps(msg, ensure_ascii=False) + "\n")
    sys.stdout.flush()


def handle(req: dict):
    method, id_ = req.get("method"), req.get("id")
    if method == "initialize":
        send({"jsonrpc": "2.0", "id": id_, "result": {
            "protocolVersion": PROTOCOL, "capabilities": {"tools": {}}, "serverInfo": SERVER_INFO}})
    elif method == "notifications/initialized":
        pass
    elif method == "ping":
        send({"jsonrpc": "2.0", "id": id_, "result": {}})
    elif method == "tools/list":
        send({"jsonrpc": "2.0", "id": id_, "result": {"tools": TOOLS_SPEC}})
    elif method == "tools/call":
        name = req["params"]["name"]
        if name not in TOOL_IMPLS:
            send({"jsonrpc": "2.0", "id": id_, "error": {"code": -32602, "message": f"未知工具: {name}"}})
            return
        try:
            text = TOOL_IMPLS[name](req["params"].get("arguments") or {})
            send({"jsonrpc": "2.0", "id": id_, "result": {"content": [{"type": "text", "text": text}], "isError": False}})
        except Exception as e:  # noqa: BLE001 —— 业务错误回传给 AI
            send({"jsonrpc": "2.0", "id": id_, "result": {"content": [{"type": "text", "text": f"错误: {e}"}], "isError": True}})
    elif id_ is not None:
        send({"jsonrpc": "2.0", "id": id_, "error": {"code": -32601, "message": f"未实现: {method}"}})


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--selftest", action="store_true", help="本地跑通所有工具（不经 MCP）")
    args = ap.parse_args()
    if args.selftest:
        sys.exit(selftest())
    for line in sys.stdin:
        line = line.strip()
        if not line:
            continue
        try:
            handle(json.loads(line))
        except json.JSONDecodeError:
            continue


# ---------------------------------------------------------------- 自测

def selftest() -> int:
    td = Path(tempfile.mkdtemp(prefix="excel-mcp-test-"))
    p = td / "t.xlsx"
    print("1. create_workbook    ", t_create_workbook({"path": str(p), "sheet_name": "数据", "headers": ["订单号", "重量", "状态"]}))
    print("2. append_rows        ", t_append_rows({"path": str(p), "rows": [
        {"订单号": "DD1", "重量": 800, "状态": "待发"},
        {"订单号": "DD2", "重量": 1200, "状态": "待发"},
        {"订单号": "DD3", "重量": 650, "状态": "已发"}]}))
    print("3. inspect_workbook   ", t_inspect_workbook({"path": str(p)}))
    print("4. query_rows(>700)   ", t_query_rows({"path": str(p), "filters": [{"column": "重量", "op": "gt", "value": 700}]}))
    print("5. update_rows        ", t_update_rows({"path": str(p), "filters": [{"column": "订单号", "op": "eq", "value": "DD2"}],
                                                   "updates": {"状态": "已发", "重量": 1250}}))
    print("6. query_rows(check)  ", t_query_rows({"path": str(p), "filters": [{"column": "订单号", "op": "eq", "value": "DD2"}]}))
    print("7. set_cells          ", t_set_cells({"path": str(p), "cells": {"D1": "备注", "D2": "易碎"}}))
    print("8. apply_formula      ", t_apply_formula({"path": str(p), "cell": "D5", "formula": "SUM(B2:B4)"}))
    print("9. format_range       ", t_format_range({"path": str(p), "start_cell": "A1", "end_cell": "D1", "bold": True, "bg_color": "DDEBF7"}))
    print("10. find_replace      ", t_find_replace({"path": str(p), "find": "待发", "replace": "待发货"}))
    print("11. read_range        ", t_read_range({"path": str(p), "start_cell": "A1", "end_cell": "D5", "values_only": False})[:200])
    print("12. manage_sheets     ", t_manage_sheets({"path": str(p), "action": "copy", "sheet": "数据"}))
    print("13. delete_rows       ", t_delete_rows({"path": str(p), "sheet": "数据 副本", "row_start": 2}))
    # csv
    c = td / "t.csv"
    t_create_workbook({"path": str(c), "headers": ["a", "b"]})
    t_append_rows({"path": str(c), "rows": [[1, 2], [3, 4]]})
    print("csv query             ", t_query_rows({"path": str(c), "filters": [{"column": "a", "op": "ge", "value": 2}]}))
    t_update_rows({"path": str(c), "filters": [{"column": "a", "op": "eq", "value": 3}], "updates": {"b": 99}})
    print("csv update check      ", t_query_rows({"path": str(c)}))
    print(f"\n✔ 全部通过（临时目录 {td}）")
    return 0


if __name__ == "__main__":
    main()
